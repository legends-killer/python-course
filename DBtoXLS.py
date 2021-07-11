import sqlite3
import xlwt
import openpyxl
import pandas as pd

class ExcelData():
    def write(self, data_path, sheetname, value):
        index = len(value)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheetname
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        workbook.save(data_path)
        print("xlsx格式表格写入数据成功！")

    def write_m2(self,datalist,savepath):
        book = xlwt.Workbook(encoding="utf-8")
        sheet = book.add_sheet('sheet1', cell_overwrite_ok=True)
        index = len(datalist)
        for i in range(0, index):
            for j in range(0, len(datalist[i])):
                sheet.write(i, j, datalist[i][j])
        book.save(savepath)

class DB():
    def __init__(self,dbname):
        self.conn = sqlite3.connect(f'./{dbname}')
        self.cursor = self.conn.cursor()
        print("Opened database successfully")

    def get_table(self):
        self.cursor.execute("select name from sqlite_master where type='table'")
        self.conn.commit()
        result = self.cursor.fetchall()
        return result

    def read(self,tablename):
        try:
            self.cursor.execute("SELECT * from " + tablename)
            self.conn.commit()
            description = self.cursor.description
            title = []

            # for data in description:
            #     title.append(data[0])
            # datas = []
            # for row in self.cursor.fetchall():
            #     sheelData = {}
            #     for col in range(len(row)):
            #         sheelData[title[col]] = row[col]
            #     datas.append(sheelData)
            for data in description:
                title.append(data[0])
            datas = []
            datas.append(title)
            for row in self.cursor.fetchall():
                sheelData = []
                for col in range(len(row)):
                    sheelData.append(row[col])
                datas.append(sheelData)
            # print(datas)
            return datas

        except Exception as e:
            print(str(e))
            print("数据读取错误")


aa = DB("content.db")
table_list = aa.get_table()
print(table_list)
for i in range(len(table_list)):
    if table_list[i][0] == 'sqlite_sequence':
        continue
    title = table_list[i][0]
    value = aa.read(title)
    ss = table_list[i][0]+".xls"
    bb = ExcelData()
    bb.write_m2(value,ss)

    data = pd.read_excel(ss, 'sheet1', index_col=0)
    data.to_csv(f'{table_list[i][0]}.csv', encoding='utf-8')
# value = aa.read("手机")
# ExcelData().write_m2(value,"phone.xls")
#ExcelData().write(value,"手机","phone.xls")
